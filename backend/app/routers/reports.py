from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func
from io import BytesIO
from datetime import datetime, date

from openpyxl import Workbook

from app.database.db import get_db
from app.database.models import Component, BorrowItem, BorrowTransaction, Container, Borrower, BorrowStatus

router = APIRouter(
    prefix="/reports",
    tags=["Reports"]
)


@router.get("/stats")
def get_dashboard_stats(db: Session = Depends(get_db)):
    # Total components in inventory
    total_components = (
        db.query(Component)
        .filter(Component.is_deleted == False)
        .count()
    )

    # Currently borrowed components (number of distinct component types being borrowed)
    currently_borrowed = (
        db.query(BorrowItem.component_id)
        .filter(BorrowItem.quantity_returned < BorrowItem.quantity_borrowed)
        .distinct()
        .count()
    )

    # Overdue borrows (count of overdue borrow transactions)
    # Overdue is only when today's date > expected_return_date (not >=)
    today = date.today()
    overdue_count = (
        db.query(BorrowItem)
        .join(BorrowItem.transaction)
        .filter(
            BorrowTransaction.expected_return_date < today,  # Use < instead of <= to match active borrow logic
            BorrowItem.quantity_returned < BorrowItem.quantity_borrowed,
            BorrowTransaction.status != BorrowStatus.COMPLETED
        )
        .distinct(BorrowTransaction.id)
        .count()
    )

    return {
        "total_components": total_components,
        "currently_borrowed": currently_borrowed,
        "overdue_borrows": overdue_count
    }


@router.get("/database/export-full")
def export_full_database(db: Session = Depends(get_db)):
    """Export entire database to Excel for visualization - includes deleted components"""
    from app.database.models import Container, Borrower
    
    wb = Workbook()
    
    # Remove default sheet
    wb.remove(wb.active)
    
    # -------------------------
    # Components Sheet (All - including deleted)
    # -------------------------
    ws_components = wb.create_sheet("Components")
    ws_components.append([
        "ID",
        "Name",
        "Category",
        "Quantity",
        "Storage Type",
        "Cabinet Number",
        "Shelf Number",
        "Container Code",
        "Drawer Index",
        "Storage Box Index",
        "Location Type",
        "Location Index",
        "Remarks",
        "Is Deleted",
        "Created At",
        "Deleted At",
        "Image Path"
    ])
    
    components = db.query(Component).order_by(Component.id).all()
    for comp in components:
        ws_components.append([
            comp.id,
            comp.name,
            comp.category,
            comp.quantity,
            comp.storage_type,
            comp.cabinet_number or "",
            comp.shelf_number or "",
            comp.container.code if comp.container else "",
            comp.drawer_index or "",
            comp.storage_box_index or "",
            comp.location_type,
            comp.location_index or "",
            comp.remarks or "",
            comp.is_deleted,
            comp.created_at.strftime("%d/%m/%Y %H:%M:%S") if comp.created_at else "",
            comp.deleted_at.strftime("%d/%m/%Y %H:%M:%S") if comp.deleted_at else "",
            comp.image_path or ""
        ])
    
    # Set column widths for Components sheet
    component_widths = {
        'A': 8,   # ID
        'B': 30,  # Name
        'C': 25,  # Category
        'D': 12,  # Quantity
        'E': 15,  # Storage Type
        'F': 15,  # Cabinet Number
        'G': 15,  # Shelf Number
        'H': 15,  # Container Code
        'I': 15,  # Drawer Index
        'J': 15,  # Storage Box Index
        'K': 15,  # Location Type
        'L': 15,  # Location Index
        'M': 35,  # Remarks
        'N': 12,  # Is Deleted
        'O': 20,  # Created At
        'P': 20,  # Deleted At
        'Q': 40   # Image Path   # Image Path
    }
    for col, width in component_widths.items():
        ws_components.column_dimensions[col].width = width
    
    # -------------------------
    # Containers Sheet
    # -------------------------
    ws_containers = wb.create_sheet("Containers")
    ws_containers.append(["ID", "Code", "Cabinet Number", "Shelf Number", "QR Path"])
    
    containers = db.query(Container).order_by(Container.id).all()
    for cont in containers:
        ws_containers.append([
            cont.id,
            cont.code,
            cont.cabinet_number,
            cont.shelf_number or "",
            cont.qr_path or ""
        ])
    
    container_widths = {'A': 8, 'B': 15, 'C': 18, 'D': 15, 'E': 40}
    for col, width in container_widths.items():
        ws_containers.column_dimensions[col].width = width
    
    # -------------------------
    # Borrow Items Sheet (Component-Borrow relationships)
    # -------------------------
    ws_borrow_items = wb.create_sheet("Borrow Items")
    ws_borrow_items.append([
        "ID",
        "Transaction ID",
        "Component ID",
        "Component Name",
        "Quantity Borrowed",
        "Quantity Returned",
        "Remaining"
    ])
    
    borrow_items = db.query(BorrowItem).order_by(BorrowItem.id).all()
    for item in borrow_items:
        ws_borrow_items.append([
            item.id,
            item.transaction_id,
            item.component_id,
            item.component.name if item.component else "N/A",
            item.quantity_borrowed,
            item.quantity_returned,
            item.quantity_borrowed - item.quantity_returned
        ])
    
    borrow_item_widths = {'A': 8, 'B': 15, 'C': 12, 'D': 30, 'E': 18, 'F': 18, 'G': 12}
    for col, width in borrow_item_widths.items():
        ws_borrow_items.column_dimensions[col].width = width
    
    # -------------------------
    # Borrow Transactions Sheet
    # -------------------------
    ws_borrow_tx = wb.create_sheet("Borrow Transactions")
    ws_borrow_tx.append([
        "ID",
        "Borrower Name",
        "TP ID",
        "Phone",
        "Email",
        "PIC Name",
        "Reason",
        "Expected Return Date",
        "Status",
        "Borrowed At"
    ])
    
    transactions = db.query(BorrowTransaction).order_by(BorrowTransaction.id).all()
    for tx in transactions:
        ws_borrow_tx.append([
            tx.id,
            tx.borrower.name if tx.borrower else "N/A",
            tx.borrower.tp_id if tx.borrower else "N/A",
            tx.borrower.phone if tx.borrower else "N/A",
            tx.borrower.email if tx.borrower else "N/A",
            tx.borrowed_by.name if tx.borrowed_by else "N/A",
            tx.reason or "",
            tx.expected_return_date.strftime("%d/%m/%Y") if tx.expected_return_date else "",
            tx.status.value if hasattr(tx.status, 'value') else str(tx.status),
            tx.borrowed_at.strftime("%d/%m/%Y %H:%M:%S") if tx.borrowed_at else ""
        ])
    
    tx_widths = {'A': 8, 'B': 25, 'C': 15, 'D': 15, 'E': 25, 'F': 20, 'G': 35, 'H': 20, 'I': 15, 'J': 20}
    for col, width in tx_widths.items():
        ws_borrow_tx.column_dimensions[col].width = width
    
    # -------------------------
    # Summary Sheet
    # -------------------------
    ws_summary = wb.create_sheet("Summary", 0)  # Insert at beginning
    total_components = db.query(Component).count()
    active_components = db.query(Component).filter(Component.is_deleted == False).count()
    deleted_components = db.query(Component).filter(Component.is_deleted == True).count()
    total_containers = db.query(Container).count()
    total_borrow_items = db.query(BorrowItem).count()
    total_transactions = db.query(BorrowTransaction).count()
    
    ws_summary.append(["Database Export Summary"])
    ws_summary.append([])
    ws_summary.append(["Total Components", total_components])
    ws_summary.append(["Active Components", active_components])
    ws_summary.append(["Deleted Components", deleted_components])
    ws_summary.append(["Total Containers", total_containers])
    ws_summary.append(["Total Borrow Items", total_borrow_items])
    ws_summary.append(["Total Borrow Transactions", total_transactions])
    ws_summary.append([])
    ws_summary.append(["Export Date", datetime.now().strftime("%d/%m/%Y %H:%M:%S")])
    
    summary_widths = {'A': 25, 'B': 20}
    for col, width in summary_widths.items():
        ws_summary.column_dimensions[col].width = width
    
    # -------------------------
    # Stream file
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    filename = f"database_export_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )


@router.get("/components/export")
def export_components_report(db: Session = Depends(get_db)):
    components = (
        db.query(Component)
        .filter(Component.is_deleted == False)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Component Report"

    # -------------------------
    # Header row
    # -------------------------
    from app.utils.component_mapper import generate_location_label
    ws.append([
        "Component Name",
        "Category",
        "Storage Type",
        "Location",
        "Total Quantity",
        "Borrowed Quantity",
        "Date Added"
    ])

    # -------------------------
    # Data rows
    # -------------------------
    for component in components:
        borrowed_qty = sum(
            (item.quantity_borrowed - item.quantity_returned)
            for item in component.borrow_items
        )
        
        location_label = generate_location_label(component)

        ws.append([
            component.name,
            component.category,
            component.storage_type,
            location_label,
            component.quantity,
            borrowed_qty,
            component.created_at.strftime("%d/%m/%Y")
        ])

    # -------------------------
    # Set column widths
    # -------------------------
    column_widths = {
        'A': 30,  # Component Name
        'B': 30,  # Category
        'C': 15,  # Storage Type
        'D': 50,  # Location
        'E': 15,  # Total Quantity
        'F': 18,  # Borrowed Quantity
        'G': 15   # Date Added
    }
    for col, width in column_widths.items():
        ws.column_dimensions[col].width = width

    # -------------------------
    # Stream file
    # -------------------------
    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = f"component_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={
            "Content-Disposition": f"attachment; filename={filename}"
        }
    )
